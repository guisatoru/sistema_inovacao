# testes/management/commands/import_gestao_pessoas.py
from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, Any, Set, Tuple, List

import openpyxl

from django.core.management.base import BaseCommand
from django.db import transaction

from testes.models import Loja, Regiao, Colaborador, Funcao


ABA_LOJAS = "Relação de lojas"
ABA_FUNC = "Relação de funcionários"


def _norm(v) -> str:
    return (str(v).strip() if v is not None else "")


def _to_int(v, default=0) -> int:
    if v is None:
        return default
    s = _norm(v)
    if not s:
        return default

    s = s.replace(",", ".")
    try:
        return int(float(s))
    except ValueError:
        return default


@dataclass
class ImportResult:
    created_regioes: int = 0
    lojas_criadas: int = 0
    lojas_atualizadas_regiao: int = 0
    lojas_reativadas: int = 0
    lojas_inativadas: int = 0
    lojas_ativas_importadas: int = 0

    colabs_criados: int = 0
    colabs_atualizados: int = 0
    colabs_inativados: int = 0
    colabs_ignorados_status: int = 0
    colabs_sem_loja_ativa: int = 0

    def to_dict(self) -> Dict[str, Any]:
        return self.__dict__


def _require_sheet(wb, name: str):
    if name not in wb.sheetnames:
        raise ValueError(f"Aba '{name}' não encontrada.")
    return wb[name]


def _header_map(ws) -> Dict[str, int]:
    header = [_norm(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    return {h: i for i, h in enumerate(header) if h}


@transaction.atomic
def run_import(filepath: str, *, dry_run: bool = False) -> Dict[str, Any]:
    """
    Import modo seguro:
    - nunca apaga
    - cria/atualiza
    - reativa quem voltou
    - desativa quem sumiu
    """
    res = ImportResult()

    # read_only melhora MUITO performance/memória pra planilha grande
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    ws_lojas = _require_sheet(wb, ABA_LOJAS)
    ws_func = _require_sheet(wb, ABA_FUNC)

    # -----------------------------
    # LOJAS (BULK)
    # -----------------------------
    h_lojas = _header_map(ws_lojas)

    def idx_lojas(name):
        if name not in h_lojas:
            raise ValueError(f"Coluna '{name}' não encontrada na aba {ABA_LOJAS}.")
        return h_lojas[name]

    idx_nome_loja = idx_lojas("LOJA")
    idx_regiao = idx_lojas("REGIÃO")
    idx_quadro = h_lojas.get("QUADRO CONTRATO")
    if idx_quadro is None:
        idx_quadro = h_lojas.get("QUADRO CONTRATADO")
    if idx_quadro is None:
        raise ValueError(f"Coluna 'QUADRO CONTRATO'/'QUADRO CONTRATADO' não encontrada na aba {ABA_LOJAS}.")

    # 1) Primeiro passa: coletar lojas válidas (quadro > 0)
    lojas_import: List[Tuple[str, str, int]] = []  # (nome_loja, regiao_nome, quadro_contratado)
    regioes_set: Set[str] = set()
    lojas_ativas_importadas: Set[str] = set()

    for row in ws_lojas.iter_rows(min_row=2, values_only=True):
        nome_loja = _norm(row[idx_nome_loja])
        regiao_nome = _norm(row[idx_regiao])
        quadro = _to_int(row[idx_quadro])

        if not nome_loja or not regiao_nome:
            continue
        if quadro <= 0:
            continue

        lojas_import.append((nome_loja, regiao_nome, quadro))
        regioes_set.add(regiao_nome)
        lojas_ativas_importadas.add(nome_loja)

    res.lojas_ativas_importadas = len(lojas_ativas_importadas)

    # 2) Garantir regiões (bulk_create das faltantes)
    regioes_exist = Regiao.objects.filter(nome__in=regioes_set).only("id", "nome")
    regiao_by_nome = {r.nome: r for r in regioes_exist}

    to_create_reg = [Regiao(nome=nome) for nome in regioes_set if nome not in regiao_by_nome]
    if to_create_reg:
        Regiao.objects.bulk_create(to_create_reg, batch_size=2000)
        res.created_regioes = len(to_create_reg)
        # recarrega as regiões criadas para ter IDs
        regioes_exist = Regiao.objects.filter(nome__in=regioes_set).only("id", "nome")
        regiao_by_nome = {r.nome: r for r in regioes_exist}

    # 3) Buscar lojas existentes e preparar create/update em lote
    lojas_exist = Loja.objects.filter(nome__in=lojas_ativas_importadas).only("id", "nome", "regiao_id", "ativo")
    loja_by_nome = {l.nome: l for l in lojas_exist}

    to_create_lojas: List[Loja] = []
    to_update_lojas: List[Loja] = []

    for nome_loja, regiao_nome, quadro in lojas_import:
        regiao_obj = regiao_by_nome.get(regiao_nome)
        if not regiao_obj:
            # Em teoria não acontece, mas mantém seguro
            continue

        loja_obj = loja_by_nome.get(nome_loja)
        if loja_obj is None:
            to_create_lojas.append(
                Loja(nome=nome_loja, regiao=regiao_obj, quadro_contratado=max(0, quadro), ativo=True)
            )
        else:
            changed = False
            if loja_obj.regiao_id != regiao_obj.id:
                loja_obj.regiao = regiao_obj
                res.lojas_atualizadas_regiao += 1
                changed = True
            if (loja_obj.quadro_contratado or 0) != max(0, quadro):
                loja_obj.quadro_contratado = max(0, quadro)
                changed = True
            if loja_obj.ativo is False:
                loja_obj.ativo = True
                res.lojas_reativadas += 1
                changed = True

            if changed:
                to_update_lojas.append(loja_obj)

    if to_create_lojas:
        Loja.objects.bulk_create(to_create_lojas, batch_size=2000)
        res.lojas_criadas = len(to_create_lojas)

    if to_update_lojas:
        # Atualiza ambos campos de uma vez (ok mesmo que só 1 tenha mudado)
        Loja.objects.bulk_update(to_update_lojas, ["regiao", "quadro_contratado", "ativo"], batch_size=2000)

    # 4) Desativar lojas que sumiram
    res.lojas_inativadas = (
        Loja.objects.filter(ativo=True)
        .exclude(nome__in=lojas_ativas_importadas)
        .update(ativo=False)
    )

    # 5) Montar lojas_map (com IDs garantidos após bulk)
    lojas_map = {
        l.nome: l
        for l in Loja.objects.filter(nome__in=lojas_ativas_importadas, ativo=True).only("id", "nome")
    }

    # -----------------------------
    # COLABORADORES (BULK)
    # -----------------------------
    h_func = _header_map(ws_func)

    def idx_func(name):
        if name not in h_func:
            raise ValueError(f"Coluna '{name}' não encontrada na aba {ABA_FUNC}.")
        return h_func[name]

    idx_re = idx_func("CÓD. FUNCIONÁRIO")
    idx_nome = idx_func("NOME COMPLETO")
    idx_loja = idx_func("LOJA")
    idx_funcao = idx_func("FUNÇÃO")
    idx_status = idx_func("STATUS")

    # 1) Coletar somente ATIVOS + validar loja
    ativos_re_importados: Set[str] = set()
    ativos_data: Dict[str, Tuple[str, Loja, str]] = {}  # re -> (nome, loja_obj, funcao_nome)
    funcoes_importadas: Set[str] = set()

    for row in ws_func.iter_rows(min_row=2, values_only=True):
        status = _norm(row[idx_status]).upper()
        if status != "ATIVO":
            res.colabs_ignorados_status += 1
            continue

        re_ = _norm(row[idx_re])
        nome = _norm(row[idx_nome])
        loja_nome = _norm(row[idx_loja])
        funcao_nome = _norm(row[idx_funcao])

        if not re_ or not nome or not loja_nome:
            continue

        loja_obj = lojas_map.get(loja_nome)
        if not loja_obj:
            res.colabs_sem_loja_ativa += 1
            continue

        ativos_re_importados.add(re_)
        ativos_data[re_] = (nome, loja_obj, funcao_nome)
        if funcao_nome:
            funcoes_importadas.add(funcao_nome)

    # 1.1) Garantir funcoes para vincular colaboradores
    if funcoes_importadas:
        funcoes_exist = Funcao.objects.filter(nome__in=funcoes_importadas).only("id", "nome")
        funcao_by_nome = {f.nome: f for f in funcoes_exist}
        to_create_funcoes = [Funcao(nome=nome) for nome in funcoes_importadas if nome not in funcao_by_nome]
        if to_create_funcoes:
            Funcao.objects.bulk_create(to_create_funcoes, batch_size=1000)
            funcoes_exist = Funcao.objects.filter(nome__in=funcoes_importadas).only("id", "nome")
            funcao_by_nome = {f.nome: f for f in funcoes_exist}
    else:
        funcao_by_nome = {}

    re_list = list(ativos_re_importados)

    # 2) Buscar existentes em 1 query (lookup O(1) com in_bulk)
    existentes = Colaborador.objects.in_bulk(re_list, field_name="re")

    to_create_colabs: List[Colaborador] = []
    to_update_colabs: List[Colaborador] = []

    for re_ in re_list:
        nome, loja_obj, funcao_nome = ativos_data[re_]
        funcao_obj = funcao_by_nome.get(funcao_nome) if funcao_nome else None
        obj = existentes.get(re_)

        if obj is None:
            to_create_colabs.append(
                Colaborador(re=re_, nome=nome, loja=loja_obj, funcao=funcao_obj, ativo=True)
            )
            res.colabs_criados += 1
        else:
            # Mantém a contagem "atualizados" parecida com update_or_create:
            # se já existia, conta como "atualizado" (mesmo que não mude nada)
            res.colabs_atualizados += 1

            changed = False
            if obj.nome != nome:
                obj.nome = nome
                changed = True
            if obj.loja_id != loja_obj.id:
                obj.loja = loja_obj
                changed = True
            if obj.funcao_id != (funcao_obj.id if funcao_obj else None):
                obj.funcao = funcao_obj
                changed = True
            if obj.ativo is False:
                obj.ativo = True
                changed = True

            if changed:
                to_update_colabs.append(obj)

    if to_create_colabs:
        Colaborador.objects.bulk_create(to_create_colabs, batch_size=5000)

    if to_update_colabs:
        Colaborador.objects.bulk_update(to_update_colabs, ["nome", "loja", "funcao", "ativo"], batch_size=5000)

    # 3) Inativar quem sumiu
    res.colabs_inativados = (
        Colaborador.objects.filter(ativo=True)
        .exclude(re__in=ativos_re_importados)
        .update(ativo=False)
    )

    if dry_run:
        transaction.set_rollback(True)

    return res.to_dict()


class Command(BaseCommand):
    help = "Importa Gestão de Pessoas (modo seguro: nunca apaga, só ativa/inativa)."

    def add_arguments(self, parser):
        parser.add_argument("arquivo", type=str)
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **opts):
        result = run_import(opts["arquivo"], dry_run=opts["dry_run"])

        self.stdout.write(self.style.SUCCESS("Import finalizado."))
        for k, v in result.items():
            self.stdout.write(f"{k}: {v}")

        if opts["dry_run"]:
            self.stdout.write(self.style.WARNING("DRY RUN — rollback aplicado"))
